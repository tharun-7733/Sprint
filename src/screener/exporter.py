"""
exporter.py — Screener Excel report exporter (Sprint 3, Day 17).

Generates output/screener_output.xlsx with one sheet per preset.
Each sheet contains 20 KPI columns, sorted by composite_quality_score,
with colour-coded cells (green = passes threshold, red = fails).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Colour constants ─────────────────────────────────────────────────────────
_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_SCORE  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 20 KPI columns in display order
_KPI_COLUMNS: list[tuple[str, str]] = [
    ("ticker",                  "Ticker"),
    ("company_name",            "Company"),
    ("sector_name",             "Sector"),
    ("composite_quality_score", "Quality Score"),
    ("sector_relative_score",   "Sector Score"),
    ("roe",                     "ROE (%)"),
    ("roce",                    "ROCE (%)"),
    ("npm",                     "NPM (%)"),
    ("operating_margin",        "OPM (%)"),
    ("debt_to_equity",          "D/E Ratio"),
    ("interest_coverage",       "ICR"),
    ("free_cashflow",           "FCF (Cr)"),
    ("revenue",                 "Revenue (Cr)"),
    ("net_profit",              "Net Profit (Cr)"),
    ("revenue_cagr_5yr",        "Rev CAGR 5yr (%)"),
    ("pat_cagr_5yr",            "PAT CAGR 5yr (%)"),
    ("eps_cagr_5yr",            "EPS CAGR 5yr (%)"),
    ("price_to_earnings",       "P/E"),
    ("price_to_book",           "P/B"),
    ("dividend_yield",          "Div Yield (%)"),
]

# Map metric name → column name in the DataFrame
_METRIC_TO_COL: dict[str, str] = {
    "roe":              "roe",
    "debt_to_equity":   "debt_to_equity",
    "free_cashflow":    "free_cashflow",
    "revenue_cagr_5yr": "revenue_cagr_5yr",
    "pat_cagr_5yr":     "pat_cagr_5yr",
    "operating_margin": "operating_margin",
    "price_to_earnings":"price_to_earnings",
    "price_to_book":    "price_to_book",
    "dividend_yield":   "dividend_yield",
    "interest_coverage":"interest_coverage",
    "market_cap":       "market_cap",
    "net_profit":       "net_profit",
    "eps_cagr_5yr":     "eps_cagr_5yr",
    "asset_turnover":   "asset_turnover",
    "revenue":          "revenue",
    "_payout_ratio_max":"payout_ratio",
}


def _cell_passes(value: Any, operator: str, threshold: float, sector: str) -> bool | None:
    """
    Determine whether a cell value passes a filter threshold.
    Returns None if the result is indeterminate (e.g., NaN).
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None

    if operator == "positive":
        return v > 0
    elif operator == "eq":
        return abs(v - threshold) < 1e-9
    elif operator == "min":
        return v >= threshold
    elif operator == "max":
        return v <= threshold
    elif operator == "declining":
        return True  # boolean already applied at filter time
    return None


def _write_sheet(
    ws: Any,
    df: pd.DataFrame,
    preset_label: str,
    filters: list[dict[str, Any]],
    config_metrics: dict[str, Any],
) -> None:
    """Write one preset sheet to openpyxl worksheet ws."""

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.append([f"Preset: {preset_label}  |  Companies: {len(df)}"])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = _HEADER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_KPI_COLUMNS))

    # ── Header row ────────────────────────────────────────────────────────────
    header_row = 2
    for col_idx, (_, label) in enumerate(_KPI_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _THIN_BORDER

    # Build a lookup: col_name → (operator, threshold, metric_key)
    filter_map: dict[str, tuple[str, float, str]] = {}
    for flt in filters:
        mkey      = flt["metric"]
        col_name  = _METRIC_TO_COL.get(mkey, mkey)
        filter_map[col_name] = (flt["operator"], float(flt.get("threshold", 0)), mkey)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        sector_name = str(row.get("sector_name", ""))
        for col_idx, (col_name, _) in enumerate(_KPI_COLUMNS, start=1):
            value = row.get(col_name, None)

            # Format numeric values
            if isinstance(value, float):
                if pd.isna(value):
                    display = "–"
                elif col_name in ("revenue", "net_profit", "free_cashflow", "market_cap"):
                    display = f"{value:,.0f}"
                elif col_name == "composite_quality_score":
                    display = round(value, 1)
                else:
                    display = round(value, 2)
            else:
                display = value

            cell = ws.cell(row=row_idx, column=col_idx, value=display)
            cell.alignment = Alignment(horizontal="right" if col_idx > 3 else "left")
            cell.border    = _THIN_BORDER
            cell.font      = Font(size=9)

            # Composite/sector score columns get blue tint
            if col_name in ("composite_quality_score", "sector_relative_score"):
                cell.fill = _SCORE
                continue

            # Colour-code based on filter threshold
            if col_name in filter_map and not isinstance(display, str):
                operator, threshold, mkey = filter_map[col_name]
                passes = _cell_passes(value, operator, threshold, sector_name)
                if passes is True:
                    cell.fill = _GREEN
                elif passes is False:
                    cell.fill = _RED

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {1: 12, 2: 28, 3: 22, 4: 13, 5: 12}
    for col_idx in range(1, len(_KPI_COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_idx, 14)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"


def export_screener_output(
    preset_results: dict[str, pd.DataFrame],
    config: dict[str, Any],
    output_path: str | Path = "output/screener_output.xlsx",
) -> Path:
    """
    Write screener_output.xlsx — one sheet per preset.

    Parameters
    ----------
    preset_results : dict mapping preset name → filtered DataFrame
    config         : full screener_config dict (for labels and filters)
    output_path    : destination .xlsx path

    Returns
    -------
    Path to the written file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    presets_cfg = config.get("presets", {})

    for preset_name, df in preset_results.items():
        label   = presets_cfg.get(preset_name, {}).get("label", preset_name)
        filters = presets_cfg.get(preset_name, {}).get("filters", [])

        # Sheet name: max 31 chars, no special chars
        sheet_name = label[:31].replace("/", "-").replace("\\", "-").replace(":", "")
        ws = wb.create_sheet(title=sheet_name)

        metric_cfg = config.get("metrics", {})
        _write_sheet(ws, df, label, filters, metric_cfg)
        logger.info("Sheet '%s' written: %d rows", sheet_name, len(df))

    wb.save(output_path)
    logger.info("screener_output.xlsx written to %s", output_path)
    return output_path
