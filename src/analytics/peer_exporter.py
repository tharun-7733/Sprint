"""
peer_exporter.py — Peer Comparison Excel Report Generator (Sprint 3, Day 20).

Generates output/peer_comparison.xlsx with one sheet per peer group.

Per sheet:
  - Columns: company_id, company_name, 20 metric columns,
             percentile rank for each of the 10 ranked metrics
  - Percentile colour-coding:
      green  ≥ 75th percentile (≥ 0.75)
      yellow  25th–75th (0.25 to 0.75)
      red    ≤ 25th percentile (≤ 0.25)
  - Benchmark company row: highlighted with gold/amber background
    (the company with the highest composite_quality_score in the group)
  - Summary row at the bottom: peer group median for each metric

Usage
-----
    from src.analytics.peer_exporter import export_peer_comparison
    from src.analytics.peer import PeerEngine
    engine = PeerEngine()
    percentile_df = engine.compute_and_persist()
    export_peer_comparison(percentile_df, "output/peer_comparison.xlsx")
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Fill colours ─────────────────────────────────────────────────────────────
_GREEN     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED       = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_GOLD      = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # benchmark
_HEADER    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_MEDIAN_BG = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─── 20 metric display columns ────────────────────────────────────────────────
_METRIC_COLS: list[tuple[str, str]] = [
    ("roe",                  "ROE (%)"),
    ("roce",                 "ROCE (%)"),
    ("npm",                  "NPM (%)"),
    ("operating_margin",     "OPM (%)"),
    ("debt_to_equity",       "D/E"),
    ("interest_coverage",    "ICR"),
    ("free_cashflow",        "FCF (Cr)"),
    ("revenue",              "Revenue (Cr)"),
    ("net_profit",           "Net Profit (Cr)"),
    ("revenue_cagr_5yr",     "Rev CAGR 5yr (%)"),
    ("pat_cagr_5yr",         "PAT CAGR 5yr (%)"),
    ("eps_cagr_5yr",         "EPS CAGR 5yr (%)"),
    ("price_to_earnings",    "P/E"),
    ("price_to_book",        "P/B"),
    ("dividend_yield",       "Div Yield (%)"),
    ("asset_turnover",       "Asset Turnover"),
    ("market_cap",           "Market Cap (Cr)"),
    ("composite_quality_score", "Quality Score"),
    ("sector_relative_score",   "Sector Score"),
    ("fcf_positive_flag",    "FCF Positive"),
]

# 10 metrics with percentile ranks (must match peer.py _RANK_METRICS keys)
_RANK_METRIC_KEYS: list[str] = [
    "ROE", "ROCE", "Net_Profit_Margin", "DE_Ratio",
    "FCF", "PAT_CAGR_5yr", "Revenue_CAGR_5yr", "EPS_CAGR_5yr",
    "Interest_Coverage", "Asset_Turnover",
]

_RANK_LABELS: dict[str, str] = {
    "ROE":              "ROE Pctile",
    "ROCE":             "ROCE Pctile",
    "Net_Profit_Margin":"NPM Pctile",
    "DE_Ratio":         "D/E Pctile",
    "FCF":              "FCF Pctile",
    "PAT_CAGR_5yr":     "PAT CAGR Pctile",
    "Revenue_CAGR_5yr": "Rev CAGR Pctile",
    "EPS_CAGR_5yr":     "EPS CAGR Pctile",
    "Interest_Coverage":"ICR Pctile",
    "Asset_Turnover":   "A.Turnover Pctile",
}


def _percentile_fill(rank: float | None) -> PatternFill | None:
    """Return the colour fill for a percentile rank cell."""
    if rank is None or (isinstance(rank, float) and pd.isna(rank)):
        return None
    if rank >= 0.75:
        return _GREEN
    elif rank <= 0.25:
        return _RED
    return _YELLOW


def _write_peer_sheet(
    ws: Any,
    group_name: str,
    universe_df: pd.DataFrame,
    percentile_df: pd.DataFrame,
) -> None:
    """Write one peer group sheet to an openpyxl worksheet."""

    # Filter universe to companies in this group
    group_pct = percentile_df[percentile_df["peer_group_name"] == group_name]
    group_ids = group_pct["company_id"].unique()
    group_co  = universe_df[universe_df["company_id"].isin(group_ids)].copy()
    group_co  = group_co.sort_values("composite_quality_score", ascending=False)

    # Benchmark: company with highest composite score
    benchmark_id = (
        group_co["company_id"].iloc[0]
        if not group_co.empty else None
    )

    # Pivot percentile data: company_id × metric → percentile_rank
    if not group_pct.empty:
        pct_pivot = group_pct.pivot_table(
            index="company_id", columns="metric",
            values="percentile_rank", aggfunc="first"
        )
    else:
        pct_pivot = pd.DataFrame()

    # ── Header rows ──────────────────────────────────────────────────────────
    col_headers = (
        ["company_id", "company_name"]
        + [label for _, label in _METRIC_COLS]
        + [_RANK_LABELS.get(m, m) for m in _RANK_METRIC_KEYS]
    )
    n_cols = len(col_headers)

    # Title
    ws.append([f"Peer Group: {group_name}  |  Companies: {len(group_co)}"])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = _HEADER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Column headers
    ws.append(col_headers)
    for c_idx in range(1, n_cols + 1):
        cell = ws.cell(row=2, column=c_idx)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = _HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _THIN

    # ── Data rows ────────────────────────────────────────────────────────────
    data_start_row = 3
    for row_offset, (_, co_row) in enumerate(group_co.iterrows()):
        ws_row = data_start_row + row_offset
        cid    = int(co_row["company_id"])
        is_benchmark = (cid == benchmark_id)

        row_values: list[Any] = [cid, co_row.get("company_name", "")]

        # Metric values
        for col_name, _ in _METRIC_COLS:
            val = co_row.get(col_name, None)
            if isinstance(val, float):
                if pd.isna(val):
                    row_values.append("–")
                elif col_name in ("revenue", "net_profit", "free_cashflow", "market_cap"):
                    row_values.append(round(val, 0))
                else:
                    row_values.append(round(val, 2))
            elif val is None:
                row_values.append("–")
            else:
                row_values.append(val)

        # Percentile rank values
        pct_row: dict[str, float | None] = {}
        if cid in pct_pivot.index:
            for mkey in _RANK_METRIC_KEYS:
                pct_row[mkey] = pct_pivot.loc[cid, mkey] if mkey in pct_pivot.columns else None
        rank_values = [pct_row.get(m, None) for m in _RANK_METRIC_KEYS]
        row_values.extend([
            f"{v:.1%}" if (v is not None and not pd.isna(v)) else "–"
            for v in rank_values
        ])

        ws.append(row_values)

        # Colour entire row gold if benchmark
        for c_idx in range(1, n_cols + 1):
            cell = ws.cell(row=ws_row, column=c_idx)
            cell.border    = _THIN
            cell.font      = Font(size=9, bold=is_benchmark)
            cell.alignment = Alignment(horizontal="right" if c_idx > 2 else "left")
            if is_benchmark:
                cell.fill = _GOLD

        # Colour percentile rank cells (cols after metrics section)
        metric_end_col = 2 + len(_METRIC_COLS)
        for rank_idx, mkey in enumerate(_RANK_METRIC_KEYS):
            rank_val = pct_row.get(mkey, None)
            c_idx    = metric_end_col + 1 + rank_idx
            cell     = ws.cell(row=ws_row, column=c_idx)
            fill     = _percentile_fill(rank_val)
            if fill and not is_benchmark:
                cell.fill = fill

    # ── Median summary row ───────────────────────────────────────────────────
    if not group_co.empty:
        median_row: list[Any] = ["", "PEER MEDIAN"]
        for col_name, _ in _METRIC_COLS:
            col_series = pd.to_numeric(group_co.get(col_name, pd.Series(dtype=float)), errors="coerce")
            med = col_series.median()
            if pd.isna(med):
                median_row.append("–")
            elif col_name in ("revenue", "net_profit", "free_cashflow", "market_cap"):
                median_row.append(round(med, 0))
            else:
                median_row.append(round(med, 2))

        # Median for percentile ranks
        if not pct_pivot.empty:
            for mkey in _RANK_METRIC_KEYS:
                if mkey in pct_pivot.columns:
                    med_pct = pct_pivot[mkey].median()
                    median_row.append(f"{med_pct:.1%}" if pd.notna(med_pct) else "–")
                else:
                    median_row.append("–")
        else:
            median_row.extend(["–"] * len(_RANK_METRIC_KEYS))

        ws.append(median_row)
        summary_ws_row = data_start_row + len(group_co)
        for c_idx in range(1, n_cols + 1):
            cell = ws.cell(row=summary_ws_row, column=c_idx)
            cell.fill      = _MEDIAN_BG
            cell.font      = Font(bold=True, size=9, italic=True)
            cell.border    = _THIN
            cell.alignment = Alignment(horizontal="right" if c_idx > 2 else "left")

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12   # company_id
    ws.column_dimensions["B"].width = 28   # company_name
    for c_idx in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 14

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "C3"


def export_peer_comparison(
    percentile_df: pd.DataFrame,
    universe_df: pd.DataFrame,
    output_path: str | Path = "output/peer_comparison.xlsx",
) -> Path:
    """
    Write peer_comparison.xlsx — one sheet per peer group.

    Parameters
    ----------
    percentile_df : output of PeerEngine.compute_and_persist() or
                    compute_peer_percentiles()
    universe_df   : full scored screener DataFrame (from ScreenerEngine)
    output_path   : destination .xlsx path

    Returns
    -------
    Path to the written file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    group_names = sorted(percentile_df["peer_group_name"].unique())
    logger.info("Writing %d peer group sheets...", len(group_names))

    wb = Workbook()
    wb.remove(wb.active)

    for group_name in group_names:
        # Sheet name: max 31 chars
        sheet_name = group_name.replace(" Peers", "")[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_peer_sheet(ws, group_name, universe_df, percentile_df)
        logger.info("  Sheet '%s' written", sheet_name)

    wb.save(output_path)
    logger.info("peer_comparison.xlsx written to %s  (%d sheets)", output_path, len(group_names))
    return output_path
